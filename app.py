"""
OI Profile Dashboard — Flask Server
  GET  /                      → public dashboard
  GET  /api/dashboard-data    → parsed JSON for dashboard
  GET  /api/pre-vs-final      → pre vs final OI comparison data
  GET  /admin                 → admin panel (password protected)
  POST /admin/upload          → upload & parse new workbook
  POST /admin/login           → authenticate
  GET  /admin/logout          → clear session
"""

import os, json, hashlib, secrets, re, struct, zlib
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime
from flask import (Flask, request, jsonify, session,
                   redirect, url_for, render_template,
                   send_from_directory, abort, Response)
from werkzeug.security import generate_password_hash
from werkzeug.utils import secure_filename
import openpyxl
from openpyxl import load_workbook

# ── CONFIG ──────────────────────────────────────────────────────────────
BASE_DIR        = os.path.dirname(os.path.abspath(__file__))
DATA_DIR        = os.path.join(BASE_DIR, 'data')
UPLOAD_DIR      = os.path.join(BASE_DIR, 'uploads')
DATA_FILE       = os.path.join(DATA_DIR, 'current.json')
PVF_FILE        = os.path.join(DATA_DIR, 'pre_vs_final.json')   # pre-vs-final store
ALLOWED         = {'.xlsx', '.xls'}
PUBLIC_SITE_URL = 'https://ratkush.pythonanywhere.com/'
OG_DESCRIPTION  = (
    'OI Profile Dashboard. Real-time tracking of Open Interest (OI) macro '
    'profiles. Monitor positioning across 7 core global markets, including '
    'SR3, Fed Funds, Treasuries, and major international rates. Analyze '
    'shifts instantly via heatmaps, historical series, or pre vs. final data '
    'views.'
)

# Markets included on the Pre vs Final page (case-insensitive substring match)
PVF_MARKET_KEYS = ['sr3', 'ff', 'treas', 'tnote', 'tbond', 'zn', 'zb', 'zt', 'zf', 'ultra']

def is_brazil_market(name):
    lc = str(name).lower()
    return 'brazil' in lc or 'brl' in lc

def is_brazil_contract(contract):
    return re.match(r'^[FJNV]\d{2}$', str(contract).strip(), re.IGNORECASE) is not None

def market_contract_cap(name):
    lc = str(name).strip().lower()
    if 'sr3' in lc:
        return 24
    if lc == 'er' or 'so3' in lc or 'sonia' in lc:
        return 20
    return None

from dotenv import load_dotenv
load_dotenv()

LOCAL_FALLBACK_HASH      = generate_password_hash('localadmin123')
ADMIN_PASSWORD_HASH      = os.environ.get("ADMIN_PASSWORD_HASH", LOCAL_FALLBACK_HASH)

app = Flask(__name__, template_folder='templates', static_folder='static')
app.secret_key = os.environ.get('FLASK_SECRET_KEY', 'dev-fallback-secret-key-999')
os.makedirs(DATA_DIR,   exist_ok=True)
os.makedirs(UPLOAD_DIR, exist_ok=True)

# ── HELPERS ──────────────────────────────────────────────────────────────
def excel_serial_to_date(serial):
    """Convert Excel date serial to dd/mm/yy string."""
    try:
        serial = float(serial)
        from datetime import date, timedelta
        base = date(1899, 12, 31)
        d = base + timedelta(days=int(serial) - 1)
        return f"{d.day:02d}/{d.month:02d}/{str(d.year)[-2:]}"
    except Exception:
        return str(serial)

def to_display_date(v):
    """Normalize any date value to dd/mm/yy."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        if 10000 < v < 80000:
            return excel_serial_to_date(v)
        return str(int(v))
    if hasattr(v, 'strftime'):
        return v.strftime('%d/%m/%y')
    s = str(v).strip()
    import re
    m = re.match(r'^(\d{4})-(\d{2})-(\d{2})$', s)
    if m:
        return f"{m.group(3)}/{m.group(2)}/{m.group(1)[-2:]}"
    return s

def parse_workbook(filepath):
    """
    Parse xlsx workbook into dashboard payload.
    Returns dict: { markets, market_order, updated_at, latest_date,
                    markets_count, contracts_count }
    """
    wb = load_workbook(filepath, data_only=True, read_only=True)
    markets      = {}
    market_order = []
    global_latest = None

    for sn in wb.sheetnames:
        if sn.endswith('_0'):
            continue
        ws  = wb[sn]
        raw = [list(row) for row in ws.iter_rows(values_only=True)]
        if not raw or len(raw) < 3:
            continue

        # Drop column B (index 1) — mirrors the JS splice(1,1)
        raw = [([r[0]] + list(r[2:])) if len(r) > 1 else list(r) for r in raw]

        # Find header row (first cell == 'dates', case-insensitive)
        header_row = -1
        for i, row in enumerate(raw):
            if row and row[0] is not None and str(row[0]).strip().lower() == 'dates':
                header_row = i
                break
        if header_row == -1:
            continue

        hdr   = raw[header_row]
        dates = []
        for c in range(1, len(hdr)):
            v = hdr[c]
            if v is None or str(v).strip() == '':
                continue
            ds = to_display_date(v)
            if ds:
                dates.append(ds)
        if not dates:
            continue

        if global_latest is None or dates[0] > global_latest:
            global_latest = dates[0]

        contracts = []
        data      = {}
        for r in range(header_row + 1, len(raw)):
            row = raw[r]
            if not row or row[0] is None or str(row[0]).strip() == '':
                continue
            c_name = str(row[0]).strip()
            if is_brazil_market(sn) and not is_brazil_contract(c_name):
                continue
            contracts.append(c_name)
            data[c_name] = {}
            for ci, date_str in enumerate(dates):
                val = row[ci + 1] if ci + 1 < len(row) else None
                if val is not None and val != '':
                    try:
                        data[c_name][date_str] = float(val)
                    except (ValueError, TypeError):
                        data[c_name][date_str] = None
                else:
                    data[c_name][date_str] = None

        if not contracts:
            continue
        cap = market_contract_cap(sn)
        if cap is not None:
            contracts = contracts[:cap]
        markets[sn]      = {'contracts': contracts, 'dates': dates, 'data': data}
        market_order.append(sn)

    wb.close()
    total_contracts = sum(len(m['contracts']) for m in markets.values())
    return {
        'markets':         markets,
        'market_order':    market_order,
        'updated_at':      datetime.utcnow().isoformat() + 'Z',
        'latest_date':     global_latest or '—',
        'markets_count':   len(markets),
        'contracts_count': total_contracts,
    }

def _is_pvf_market(market_name):
    """Return True if this market should appear on the Pre vs Final page."""
    lc = market_name.lower()
    return any(k in lc for k in PVF_MARKET_KEYS)

def snapshot_pre_oi(current_payload):
    """
    Extract first-two-column snapshot from current data for PvF markets.
    Stores: { market_name: { contract: latest_oi, ... }, ... }
    """
    if not current_payload:
        return None
    snap = {
        'snapshot_date': current_payload.get('latest_date', '—'),
        'updated_at':    current_payload.get('updated_at', '—'),
        'markets':       {},
        'market_order':  [],
    }
    for mkt_name in current_payload.get('market_order', []):
        if not _is_pvf_market(mkt_name):
            continue
        mkt   = current_payload['markets'][mkt_name]
        dates = mkt.get('dates', [])
        if not dates:
            continue
        latest_date = dates[0]           # dates[0] is most recent
        contracts   = mkt.get('contracts', [])
        oi_map      = {}
        for c in contracts:
            val = mkt['data'].get(c, {}).get(latest_date)
            oi_map[c] = val
        snap['markets'][mkt_name]   = oi_map
        snap['market_order'].append(mkt_name)
    return snap

def load_current():
    if os.path.exists(DATA_FILE):
        with open(DATA_FILE, 'r') as f:
            return json.load(f)
    return None

def save_current(payload):
    with open(DATA_FILE, 'w') as f:
        json.dump(payload, f)

def load_pvf():
    if os.path.exists(PVF_FILE):
        with open(PVF_FILE, 'r') as f:
            return json.load(f)
    return None

def save_pvf(pvf):
    with open(PVF_FILE, 'w') as f:
        json.dump(pvf, f)

def admin_required(fn):
    from functools import wraps
    @wraps(fn)
    def wrapper(*args, **kwargs):
        if not session.get('admin'):
            return redirect(url_for('admin_login'))
        return fn(*args, **kwargs)
    return wrapper

def _latest_date_for_title(latest_date):
    """Convert dd/mm/yy to dd Mon for share titles."""
    if not latest_date:
        return None
    s = str(latest_date).strip()
    try:
        return datetime.strptime(s, '%d/%m/%y').strftime('%d %b')
    except ValueError:
        return s if s and s != '—' else None

def build_share_meta():
    data = load_current() or {}
    date_label = _latest_date_for_title(data.get('latest_date'))
    title = f"OI Profile {date_label}" if date_label else 'OI Profile Dashboard'
    return {
        'title': title,
        'description': OG_DESCRIPTION,
        'url': PUBLIC_SITE_URL,
        'image': PUBLIC_SITE_URL.rstrip('/') + url_for('og_image'),
    }

def _png_chunk(kind, data):
    return (
        struct.pack('>I', len(data)) +
        kind +
        data +
        struct.pack('>I', zlib.crc32(kind + data) & 0xffffffff)
    )

def _dashboard_preview_png(width=1200, height=630):
    bg = (16, 22, 31)
    pixels = bytearray(bg * width * height)

    def put(x, y, color):
        if 0 <= x < width and 0 <= y < height:
            i = (y * width + x) * 3
            pixels[i:i + 3] = bytes(color)

    def rect(x, y, w, h, color):
        x0, x1 = max(0, x), min(width, x + w)
        y0, y1 = max(0, y), min(height, y + h)
        row = bytes(color) * max(0, x1 - x0)
        for yy in range(y0, y1):
            i = (yy * width + x0) * 3
            pixels[i:i + len(row)] = row

    def line(x0, y0, x1, y1, color):
        dx = abs(x1 - x0)
        sx = 1 if x0 < x1 else -1
        dy = -abs(y1 - y0)
        sy = 1 if y0 < y1 else -1
        err = dx + dy
        while True:
            for ox in range(-1, 2):
                for oy in range(-1, 2):
                    put(x0 + ox, y0 + oy, color)
            if x0 == x1 and y0 == y1:
                break
            e2 = 2 * err
            if e2 >= dy:
                err += dy
                x0 += sx
            if e2 <= dx:
                err += dx
                y0 += sy

    rect(0, 0, width, 76, (20, 27, 40))
    rect(0, 75, width, 2, (46, 63, 88))
    rect(36, 24, 160, 18, (77, 184, 255))
    rect(232, 18, 84, 12, (122, 149, 184))
    rect(232, 38, 136, 18, (220, 232, 248))
    rect(404, 18, 104, 12, (122, 149, 184))
    rect(404, 38, 92, 18, (220, 232, 248))
    rect(1022, 22, 118, 34, (24, 31, 45))
    rect(1148, 35, 14, 14, (80, 232, 144))

    rect(0, 76, width, 54, (24, 31, 45))
    for i, x in enumerate([34, 154, 274, 394, 514, 634, 754]):
        rect(x, 98, 82, 16, (77, 184, 255) if i == 0 else (154, 176, 204))
    rect(28, 128, 112, 2, (77, 184, 255))

    rect(36, 168, 548, 390, (24, 31, 45))
    rect(36, 168, 548, 2, (58, 80, 112))
    rect(62, 198, 146, 18, (220, 232, 248))
    for i in range(9):
        rect(62 + i * 54, 242, 42, 14, (122, 149, 184))
    colors = [
        (80, 232, 144), (77, 184, 255), (255, 212, 77),
        (255, 107, 107), (176, 136, 255), (64, 221, 176)
    ]
    for r in range(9):
        rect(62, 282 + r * 26, 92, 12, (220, 232, 248))
        for c in range(7):
            color = colors[(r + c) % len(colors)]
            rect(182 + c * 50, 278 + r * 26, 38, 18, color)

    rect(622, 168, 542, 390, (24, 31, 45))
    rect(622, 168, 542, 2, (58, 80, 112))
    rect(648, 198, 190, 18, (220, 232, 248))
    for i in range(6):
        y = 262 + i * 42
        rect(660, y, 450, 1, (46, 63, 88))
    pts = [(668, 432), (744, 386), (820, 408), (896, 318), (972, 344), (1048, 254), (1114, 286)]
    for a, b in zip(pts, pts[1:]):
        line(a[0], a[1], b[0], b[1], (77, 184, 255))
    for x, y in pts:
        rect(x - 5, y - 5, 10, 10, (80, 232, 144))

    raw = b''.join(b'\x00' + pixels[y * width * 3:(y + 1) * width * 3] for y in range(height))
    png = b'\x89PNG\r\n\x1a\n'
    png += _png_chunk(b'IHDR', struct.pack('>IIBBBBB', width, height, 8, 2, 0, 0, 0))
    png += _png_chunk(b'IDAT', zlib.compress(raw, 9))
    png += _png_chunk(b'IEND', b'')
    return png

# ── PUBLIC ROUTES ────────────────────────────────────────────────────────
@app.route('/')
def index():
    return render_template('dashboard.html', share_meta=build_share_meta())

@app.route('/og-image.png')
def og_image():
    resp = Response(_dashboard_preview_png(), mimetype='image/png')
    resp.headers['Cache-Control'] = 'public, max-age=3600'
    return resp

@app.route('/sitemap.xml')
def sitemap():
    xml = f"""<?xml version="1.0" encoding="UTF-8"?>
<urlset xmlns="http://www.sitemaps.org/schemas/sitemap/0.9">
  <url>
    <loc>{PUBLIC_SITE_URL}</loc>
  </url>
</urlset>
"""
    return Response(xml, mimetype='application/xml')

@app.route('/api/dashboard-data')
def api_dashboard_data():
    data = load_current()
    if data is None:
        return jsonify({'error': 'No data published yet'}), 404
    return jsonify(data)

@app.route('/api/pre-vs-final')
def api_pre_vs_final():
    """
    Returns comparison payload:
    {
      pre:  { snapshot_date, markets: { mkt: { contract: oi } } },
      final: { snapshot_date, markets: { mkt: { contract: oi } } },
      market_order: [...],
    }
    Both pre and final are snapshots of the latest-column OI,
    taken at different publish times.
    """
    pvf     = load_pvf()
    current = load_current()
    if pvf is None or current is None:
        return jsonify({'error': 'No Pre vs Final data available yet'}), 404

    # Build 'final' snapshot from current data (same structure as pre)
    final_snap = snapshot_pre_oi(current)
    if final_snap is None:
        return jsonify({'error': 'Could not build final snapshot'}), 500

    # Unified market order: union of both, PvF markets only
    seen  = set()
    order = []
    for m in (pvf.get('market_order', []) + final_snap.get('market_order', [])):
        if m not in seen:
            seen.add(m)
            order.append(m)

    return jsonify({
        'pre':          pvf,
        'final':        final_snap,
        'market_order': order,
    })

# ── ADMIN ROUTES ─────────────────────────────────────────────────────────
@app.route('/admin', methods=['GET'])
@admin_required
def admin_panel():
    data = load_current()
    meta = {}
    pvf  = load_pvf()
    if data:
        meta = {
            'updated_at':      data.get('updated_at', '—'),
            'latest_date':     data.get('latest_date', '—'),
            'markets_count':   data.get('markets_count', 0),
            'contracts_count': data.get('contracts_count', 0),
        }
    pvf_meta = {}
    if pvf:
        pvf_meta = {
            'snapshot_date': pvf.get('snapshot_date', '—'),
            'updated_at':    pvf.get('updated_at', '—'),
            'markets':       pvf.get('market_order', []),
        }
    return render_template('admin.html', meta=meta, pvf_meta=pvf_meta)

@app.route('/admin/login', methods=['GET', 'POST'])
def admin_login():
    error = None
    if request.method == 'POST':
        pw = request.form.get('password', '')
        if check_password_hash(ADMIN_PASSWORD_HASH, pw):
            session['admin'] = True
            return redirect(url_for('admin_panel'))
        error = 'Invalid password.'
    return render_template('login.html', error=error)

@app.route('/admin/logout')
def admin_logout():
    session.clear()
    return redirect(url_for('admin_login'))

@app.route('/admin/upload', methods=['POST'])
@admin_required
def admin_upload():
    f = request.files.get('file')
    if not f or not f.filename:
        return jsonify({'error': 'No file provided'}), 400
    ext = os.path.splitext(f.filename)[1].lower()
    if ext not in ALLOWED:
        return jsonify({'error': 'Only .xlsx/.xls files accepted'}), 400

    fname = secure_filename(f.filename)
    fpath = os.path.join(UPLOAD_DIR, fname)
    f.save(fpath)

    try:
        parsed = parse_workbook(fpath)
    except Exception as e:
        return jsonify({'error': f'Parse failed: {str(e)}'}), 500

    return jsonify({
        'preview': {
            'markets_count':   parsed['markets_count'],
            'contracts_count': parsed['contracts_count'],
            'latest_date':     parsed['latest_date'],
            'markets':         parsed['market_order'],
        },
        'temp_file': fname,
    })

@app.route('/admin/publish', methods=['POST'])
@admin_required
def admin_publish():
    body      = request.json if request.is_json else request.form
    fname     = body.get('temp_file')
    is_final  = str(body.get('is_final', 'false')).lower() in ('true', '1', 'yes')

    if not fname:
        return jsonify({'error': 'No temp_file specified'}), 400
    fpath = os.path.join(UPLOAD_DIR, secure_filename(fname))
    if not os.path.exists(fpath):
        return jsonify({'error': 'File not found'}), 404

    try:
        parsed = parse_workbook(fpath)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

    # If this is a Final OI upload:
    #   1. Snapshot current data's latest-column into pre_vs_final.json BEFORE overwriting
    #   2. Then save new data as current
    if is_final:
        current = load_current()
        if current:
            snap = snapshot_pre_oi(current)
            if snap:
                save_pvf(snap)

    parsed['is_final'] = is_final   # persisted into current.json
    save_current(parsed)
    return jsonify({
        'ok':         True,
        'updated_at': parsed['updated_at'],
        'is_final':   is_final,
    })

if __name__ == '__main__':
    port  = int(os.environ.get("PORT", 5001))
    debug = os.environ.get("FLASK_DEBUG", "true").lower() == "true"
    app.run(host='0.0.0.0', port=port, debug=debug)
